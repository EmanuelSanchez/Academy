#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import os
import shutil
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook

def load_configuration ( ):
	try:
		f = open('confg.txt', 'r')
		try:
			name_doc = f.readline()
			workbook = load_workbook(filename = name_doc)
			f.close()
			return workbook
		except:
			messagebox.showerror('Error', 'No se pudo abrir el archivo seleccionado anteriormente.\nSeleccione un nuevo archivo para realizar la busqueda.')
			ejecutar (load_file())
			f = open('confg.txt', 'r')
			print("\n\n2" + "\n\n")
			workbook = load_workbook(filename = f.readline())
			f.close()
			return workbook
	except:
		workbook = error_file()
		return workbook

def save_configuration ( name_doc, addres_doc ):
	f1 = open('confg.txt', 'w')
	f1.write(name_doc)
	f1.close
	f2 = open('addres.txt', 'w')
	f2.write(addres_doc)
	f2.close()

def load_file():
  addres_doc = filedialog.askopenfilename(filetypes=[('Archivos Excel', 'xlsx')], initialdir='C:\\', parent=v0, title='Buscar Archivo')
  if addres_doc:
  	addres_act = os.getcwd( )
  	copy_doc( addres_doc, addres_act )
  	name_doc = os.path.basename( addres_doc )
  	save_configuration ( name_doc, addres_doc )
  	messagebox.showinfo('Nota:', 'Archivo importado satisfactoriamente')
  	e0.focus()
  else:
  		messagebox.showwarning('Cuidado.', 'Ningún archivo ha sido seleccionado.')
  		e0.focus()

def load_newfile():
	load_file()
	global workbook 
	workbook = load_configuration()

def update_file():
	try:
		f1 = open ('confg.txt', 'r')
		name_doc = f1.readline()
		f1.close()
		print(name_doc)
		f2 = open('addres.txt', 'r')
		addres_doc = f2.readline()
		f2.close()
		print(addres_doc)
		addres_act = os.getcwd( )
		print (addres_act)
		copy_doc( addres_doc, addres_act )
		global workbook
		workbook = load_workbook(filename = name_doc)
		f2.close()
		messagebox.showinfo('Nota', 'Archivo actualizado correctamente.')
		e0.focus()
	except:
		messagebox.showerror('Error', 'No se puedo actualizar el archivo.')
		e0.focus()

def error_file ():
	if messagebox.askokcancel("Ningún archivo ha sido seleccionado.", "No se ha seleccionado ningun archivo para realizar la búsqueda.\n¿Desea seleccionar un archivo?\n\tOk: para seleccionar el archivo.\n\tCancel: para cerrar el programa."):
		ejecutar ( load_file() )
		worbook = load_configuration()
		return worbook
	else:
		if messagebox.askokcancel("Verificación", "¿Verdaderamente desea salir?"):
			v0.withdraw()
			exit()
		else:
			error_file ( )

def copy_doc ( src, dest ):
	shutil.copy(src, dest)

def mostrar ( ventana ) :
	ventana.deiconify()

def ocultar ( ventana ) : 
	ventana.withdraw ()

def ejecutar ( f ) : 
	v0.after ( 200, f )

def Get_Num_Rows ( worksheet ):
	sheet_ranges = worksheet.calculate_dimension()
	if len(sheet_ranges) == 7:
		num_rows =  sheet_ranges[-3] + sheet_ranges[-2] + sheet_ranges[-1]
	elif len(sheet_ranges) == 6:
		num_rows =  sheet_ranges[-2] + sheet_ranges[-1]
	elif len(sheet_ranges) == 5:
		num_rows =  sheet_ranges[-1]
	else:
		messagebox.showerror("Error:","Rango de la hoja no valido")
	return (int(num_rows))

def Print_Information( worsheet, row_id ):
	ejecutar(mostrar(v1))
	information = Text( v1, height = 5, width = 40, bg = 'black', bd = 5, fg = 'white', relief = RAISED, spacing1 = 6 )
	information.insert(INSERT, 'Nombre: ')
	information.insert(INSERT, worsheet.cell(row = row_id, column = 2).value )
	information.insert(INSERT, '\nTeléfono: ')
	information.insert(INSERT, worsheet.cell(row = row_id, column = 4).value )
	information.insert(INSERT, '\nRed: ')
	information.insert(INSERT, worsheet.cell(row = row_id, column = 6).value )
	information.insert(INSERT, '\nFecha de Registro: ')
	information.insert(INSERT, worsheet.cell(row = row_id, column = 7).value )
	information.insert(INSERT, '\nNota Final: ')
	information.insert(INSERT, worsheet.cell(row = row_id, column = 9).value )
	information.grid ( column = 1 , row = 1)

def Print_Error ( Error):		
	if Error == 1:
		Error = 0
		ejecutar(mostrar(v1))
		information = Text( v1, height = 5, width = 40, bg = 'black', bd = 5, fg = 'white', relief = RAISED, spacing1 = 6 )
		information.insert(INSERT, '\n\nNo se encuentra en la base de datos.' )
		information.grid ( column = 1 , row = 1)

	elif Error == 2:
		Error = 0
		ejecutar(mostrar(v1))
		information = Text( v1, height = 5, width = 40, bg = 'black', bd = 5, fg = 'white', relief = RAISED, spacing1 = 6 )
		information.insert(INSERT, '\n\nSe encuentra registrado múltiples veces.' )
		information.grid ( column = 1 , row = 1)

def Search_ID ( num_id ):
	i = 0
	found = 0
	for sheet in workbook.get_sheet_names():
		worksheet = workbook.worksheets[i]
		i+=1
		num_rows = Get_Num_Rows( worksheet )

		for j in range(1, num_rows):
			if num_id == str(worksheet.cell(row = j, column = 3).value):
				found += 1
				row_id = j
				break
			else:
				found += 0
		if found != 0:
			break
	if found > 1:
		Error = 2
		Print_Error( Error )
	elif found == 1:
		Print_Information( worksheet, row_id )
	else:
		Error = 1
		Print_Error( Error )

def Search ():
	if e0.get() == '':
		messagebox.showerror('Error', 'No se insertó un número de cédula')
		e0.focus()
	else:
		_id = e0.get()
		Search_ID( _id )

v0 = Tk()
v1 = Toplevel(v0)

v0.config ( bg = "white" )
v0.geometry ( "320x200" )
v0.title ( "Academia de Líderes" )

v1.config ( bg = "black" )
v1.geometry ( "333x150" )
v1.title ( "Resultado" )

mainframe = ttk.Frame( v0)
mainframe.__init__(v0, padding = '4 4 50 50')
mainframe.grid( column = 0, row = 0, sticky = (N, W, E, S) )
mainframe.columnconfigure( 0, weight = 1 ) 
mainframe.rowconfigure( 0, weight = 1 )

b0 = ttk.Button(mainframe, text="Buscar", command=lambda: ejecutar ( Search( ) ))
b0.grid(column=2, row=3, sticky=E)

b1 = ttk.Button ( v1, text = "Cerrar", command = lambda: ejecutar ( ocultar(v1)) ) # Primer boton
b1.grid ( column = 1, row = 2)
v1.withdraw()

membrete = Label ( mainframe, text = "APOSENTO ALTO\nACADEMIA DE LÍDERES", relief = RAISED )
membrete.grid ( column = 2 , row = 1)
membrete.config ( bg = "red", fg = "black", relief = "solid", width = "20", height = "3") 

logoAA = PhotoImage (file = os.getcwd()+"\logoAA.gif" )
labe0 = Label( mainframe, image = logoAA)
labe0.grid ( column = 1, row = 1, padx = 10, pady = 5 )

e0 = Entry( mainframe, width = 10, textvariable = "Cédula", exportselection = 0, fg = 'black', bd = 3 )
e0.grid ( column = 1, row = 2, padx = 2, pady = 10 )

label1 = Label ( mainframe , text = "Intruduza un número de cédula ")
label1.grid ( column = 2, row = 2, padx = 2,  pady = 5)

menu = Menu(mainframe)
filemenu = Menu(menu, tearoff=0)
v0.config(menu=menu)
menu.add_cascade(label="Archivo", menu=filemenu)
filemenu.add_command(label="Actualizar archivo", command=update_file)
filemenu.add_command(label="Importar un nuevo archivo", command=load_newfile)
filemenu.add_separator()
filemenu.add_command(label="Salir", command=v0.quit)

workbook = load_configuration()

e0.focus()

v0.mainloop()	